from openpyxl import load_workbook
import os
from openpyxl.utils import get_column_letter
import string


# Excel klasörünün içerisindeki AnkaraUni Excel dosyası okunuyor.
dosya = "{}\Excel\AnkaraUni.xlsx".format(os.getcwd()) 
wb = load_workbook(filename = dosya)
sheet_ranges = wb['AnkaraUni']

# Excel'deki kutucuğun dolu olup olmadığı kontrol ediliyor.
def deger_oku(deger):
    if sheet_ranges[deger + '2'].value != None:
        return True

# Belirli sütun ve satır sayısı verildikten sonra o sütundaki dersler çekiliyor.
def ders_listesi_olustur(sutun, satir):
    ders_listesi = []
    while satir != 75:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            ders_listesi.append(sheet_ranges[saat].value)
        satir += 4
    return ders_listesi

# Belirli sütun ve satır sayısı verildikten sonra o sütundaki sınıf ve kapasiteler çekiliyor.
def sinif_ve_kapasite_listesi_olustur(sutun):
    ders_listesi = []
    sinif_listesi = sinif_listesi_olustur()
    sinif_kapasite_listesi = sinif_kapasitesi_olustur()
    z = 0
    satir = 3
    while satir != 75:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            ders_listesi.append(sheet_ranges[saat].value + '*' + sinif_listesi[z] + '*' + str(sinif_kapasite_listesi[z]))
        satir += 4
        z += 1
    return ders_listesi

# Belirli sütun ve satır sayısı verildikten sonra o sütundaki ders isimleri çekiliyor.
def ders_listesi_bos_kontrol(sutun, satir):
    ders_listesi = []
    while satir != 75:
        saat = sutun + str(satir)
        ders_listesi.append(sheet_ranges[saat].value)
        satir += 4
    return ders_listesi

# Belirli sütun ve satır sayısı verildikten sonra o sütundaki sınıf isimleri çekiliyor.
def sinif_listesi_olustur():
    sinif_listesi = []
    sutun = "A"
    satir = 3
    while satir != 75:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            sinif_listesi.append(sheet_ranges[saat].value)
        satir += 4
    return sinif_listesi

# Excel dosyasındaki tüm günler çekiliyor.
def izinliler_icin_gunler_listesi():
    gunler = []
    alfabe = list(string.ascii_uppercase)
    for harf in alfabe:
        if sheet_ranges[harf + "1"].value != None:
            gunler.append(str(sheet_ranges[harf + "1"].value).split(" ")[0])
    for harf in alfabe:
        if sheet_ranges["A{}".format(harf) + "1"].value != None:
            gunler.append(str(sheet_ranges["A{}".format(harf) + "1"].value).split(" ")[0])
    for harf in alfabe:
        if sheet_ranges["B{}".format(harf) + "1"].value != None:
            gunler.append(str(sheet_ranges["B{}".format(harf) + "1"].value).split(" ")[0])
    for harf in alfabe:
        if sheet_ranges["C{}".format(harf) + "1"].value != None:
            gunler.append(str(sheet_ranges["C{}".format(harf) + "1"].value).split(" ")[0])
    return gunler

# B sütunundaki tüm sınıf kapasiteleri çekiliyor.
def sinif_kapasitesi_olustur():
    sinif_listesi = []
    sutun = "B"
    satir = 3
    while satir != 75:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            sinif_listesi.append(sheet_ranges[saat].value)
        satir += 4
    return sinif_listesi

# Verilen sütundaki her sınavın öğrenci sayısı çekiliyor.
def ogrenci_sayisi_listesi_olustur(sutun, satir):
    ogrenci_sayisi = []
    while satir != 76:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            ogrenci_sayisi.append(sheet_ranges[saat].value)
        satir += 4
    return ogrenci_sayisi

# Verilen sütundaki her dersin sahibi çekiliyor.
def ogretmen_listesi_olustur(sutun, satir):
    ogretmen_listesi = []
    while satir != 77:
        saat = sutun + str(satir)
        if sheet_ranges[saat].value != None:
            ogretmen_listesi.append(sheet_ranges[saat].value)
        satir += 4
    return ogretmen_listesi

# Verilen sütun ve satırdaki tarih çekiliyor.
def tarih(sutun, satir):
    saat = sutun + str(satir)
    if sheet_ranges[saat].value != None:
        return sheet_ranges[saat].value

# Kullanılmıyor yine de hata almamak için silmedim.
def gunu_bul(sutun):
    a = 0
    geri_git = sheet_ranges["{}2".format(sutun)].value.hour - 9
    gun_tablosu = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    while True:
        if sheet_ranges["{}2".format(sutun)].value.hour == 9:
             tarih = sheet_ranges["{}1".format(sutun)].value
        else:
            a = 1
            while True:
                if sutun == get_column_letter(a):
                    yeni_sutun = get_column_letter(a - geri_git)
                    tarih = sheet_ranges["{}1".format(yeni_sutun)].value
                    break
                a += 1
        return str(tarih).split(" ")[0] + " " + gun_tablosu[tarih.weekday()]


# Verilen sütundaki tarih güne çevriliyor ve dönderiliyor.
def izin_icin_gunu_bul(sutun):
    a = 0
    geri_git = sheet_ranges["{}2".format(sutun)].value.hour - 9
    gun_tablosu = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    while True:
        if sheet_ranges["{}2".format(sutun)].value.hour == 9:
             tarih = sheet_ranges["{}1".format(sutun)].value
        else:
            a = 1
            while True:
                if sutun == get_column_letter(a):
                    yeni_sutun = get_column_letter(a - geri_git)
                    tarih = sheet_ranges["{}1".format(yeni_sutun)].value
                    break
                a += 1
        return str(tarih).split(" ")[0]

# Gelen öğretmen listesindeki birden fazla olan öğretmen sayısı SQL'deki DISTINCT mantığı ile teke düşürülüyor.
def gozetmen_sinav_sayisi(ogretmen_listesi):
    yeni_liste = []
    yeni_sozluk = {}
    indis  = 0
    for i in ogretmen_listesi:
        if i not in yeni_liste:
            yeni_liste.append(i)
        else:
            x = 0
            for a in yeni_liste:
                if i == a:
                    yeni_liste.pop(x)
                x += 1
            yeni_sozluk[i] = indis
        indis  += 1

    a = 0
    for i in yeni_liste:
        if i in yeni_sozluk:
            yeni_liste.pop(a)
        a += 1
            
    yeni_liste.append(yeni_sozluk)
    
    return yeni_liste